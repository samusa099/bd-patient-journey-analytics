from pathlib import Path
import csv, json, random, shutil, math
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

ROOT=Path(__file__).resolve().parents[1]
random.seed(20260821)

DIRS=[
'data/01_raw_synthetic','data/02_identity_resolution','data/03_workflow','data/04_routing',
'data/05_billing_reports','data/06_hr_operations','data/07_financial_model','data/08_predictive_operations',
'data/09_analysis_ready','data_dictionary','excel','notebooks','src','assignments/01_beginner_excel',
'assignments/02_intermediate_python','assignments/03_advanced_optimization','assignments/04_hrbp_business_case',
'case_study','competition','releases','kaggle','.github/ISSUE_TEMPLATE','.github/DISCUSSION_TEMPLATE'
]
for d in DIRS:(ROOT/d).mkdir(parents=True,exist_ok=True)

def write_csv(rel,rows):
    p=ROOT/rel;p.parent.mkdir(parents=True,exist_ok=True)
    with p.open('w',newline='',encoding='utf-8-sig') as f:
        w=csv.DictWriter(f,fieldnames=list(rows[0]));w.writeheader();w.writerows(rows)

def norm(s):
    repl={'mohammad':'md','muhammad':'md','mohammed':'md','ahmed':'ahmad'}
    return ' '.join(repl.get(x,x) for x in s.lower().replace('.','').replace('-',' ').split())

def nb(title,desc,code):
    return {'cells':[{'cell_type':'markdown','metadata':{},'source':[f'# {title}\n',desc]},
                     {'cell_type':'code','execution_count':None,'metadata':{},'outputs':[],'source':[code]}],
            'metadata':{'kernelspec':{'display_name':'Python 3','language':'python','name':'python3'}},'nbformat':4,'nbformat_minor':5}

# Hospital graph
locs=[('Reception','Main',1,'Registration',8),('Medicine OPD','Main',3,'Consultation',12),('Cardiology','Main',5,'Consultation',8),('Lab A','Diagnostic',2,'Laboratory',6),('Lab B','Diagnostic',4,'Laboratory',5),('Radiology A','Diagnostic',3,'Imaging',4),('Radiology B','Diagnostic',6,'Imaging',3),('ECG','Main',5,'Cardiac Test',4),('Ultrasound','Diagnostic',7,'Imaging',3),('Central Billing','Main',1,'Billing',6),('Report Hub','Main',1,'Reports',5),('Pharmacy','Main',1,'Pharmacy',6),('Exit','Main',1,'Exit',99)]
nodes=[{'node_id':f'N{i:02d}','node_name':n,'building':b,'floor':f,'service_type':s,'base_capacity_per_hour':c} for i,(n,b,f,s,c) in enumerate(locs,1)]
nid={r['node_name']:r['node_id'] for r in nodes}
ep=[('Reception','Medicine OPD',180),('Reception','Cardiology',220),('Medicine OPD','Lab A',220),('Medicine OPD','Lab B',310),('Medicine OPD','Radiology A',260),('Medicine OPD','ECG',180),('Medicine OPD','Ultrasound',420),('Cardiology','Lab A',260),('Cardiology','Lab B',210),('Cardiology','Radiology A',240),('Cardiology','ECG',80),('Cardiology','Ultrasound',330),('Lab A','Lab B',160),('Lab A','Radiology A',120),('Lab A','ECG',210),('Lab A','Ultrasound',280),('Lab B','Radiology B',110),('Lab B','ECG',130),('Lab B','Ultrasound',180),('Radiology A','Radiology B',210),('Radiology A','ECG',150),('Radiology B','Ultrasound',90),('ECG','Ultrasound',170),('Lab A','Central Billing',260),('Lab B','Central Billing',330),('Radiology A','Central Billing',280),('Radiology B','Central Billing',390),('ECG','Central Billing',210),('Ultrasound','Central Billing',420),('Central Billing','Report Hub',60),('Report Hub','Pharmacy',70),('Pharmacy','Exit',45),('Report Hub','Exit',40)]
edges=[{'edge_id':f'E{i:03d}','from_node':nid[a],'to_node':nid[b],'from_name':a,'to_name':b,'distance_m':d,'walk_min':round(d/75,2),'bidirectional':True} for i,(a,b,d) in enumerate(ep,1)]
g=defaultdict(list)
for e in edges:g[e['from_name']].append((e['to_name'],e['distance_m']));g[e['to_name']].append((e['from_name'],e['distance_m']))
def dist(a,b):
    import heapq
    q=[(0,a)];seen=set()
    while q:
        c,n=heapq.heappop(q)
        if n in seen:continue
        if n==b:return c
        seen.add(n)
        for x,w in g[n]:heapq.heappush(q,(c+w,x))
    return 9999

services={'CBC':('Laboratory',700,12),'HbA1c':('Laboratory',1200,14),'Blood Glucose':('Laboratory',450,8),'Lipid Profile':('Laboratory',1400,15),'Chest X-Ray':('Imaging',1200,14),'ECG':('Cardiac Test',900,10),'Ultrasound Abdomen':('Imaging',2200,20),'Creatinine':('Laboratory',650,9),'TSH':('Laboratory',1500,12)}
sloc={'Laboratory':['Lab A','Lab B'],'Imaging':['Radiology A','Radiology B'],'Cardiac Test':['ECG']}

# Synthetic identities
first=['Mohammad','Muhammad','Md','Ahmed','Ahmad','Siam','Nusrat','Fatema','Farhana','Ayesha','Rahim','Karim','Hasan','Hossain','Sadia','Mim','Rafi','Jannat']
last=['Rahman','Hossain','Ahmed','Khan','Islam','Akter','Begum','Hasan','Miah','Sarker','Chowdhury','Alam']
district=['Dhaka','Natore','Rajshahi','Chattogram','Cumilla','Gazipur','Bogura','Khulna']
patients=[]
for i in range(1,2501):
    name=f'{random.choice(first)} {random.choice(last)}';dob=datetime(1945,1,1)+timedelta(days=random.randint(0,80*365));phone=f'01{random.randint(3,9)}{random.randint(10000000,99999999)}'
    patients.append({'patient_id':f'P{i:06d}','canonical_name':name,'dob':dob.date().isoformat(),'sex':random.choice(['Female','Male']),'phone_last4':phone[-4:],'district':random.choice(district),'preferred_identity_token':random.choices(['QR','NFC','Mobile Token','Manual ID'],[.34,.32,.22,.12])[0],'synthetic':True})
variants=[];searches=[];sid=0
for p in patients:
    vs={p['canonical_name']};n=p['canonical_name']
    if n.startswith(('Mohammad ','Muhammad ')):vs.add('Md '+n.split(' ',1)[1])
    if n.startswith('Ahmed '):vs.add('Ahmad '+n.split(' ',1)[1])
    if random.random()<.35:vs.add(n.lower())
    if random.random()<.22:vs.add(n[0]+'. '+n.split()[-1])
    for v in sorted(vs):variants.append({'patient_id':p['patient_id'],'variant_name':v,'variant_type':'canonical' if v==n else 'synthetic_variant'})
    for _ in range(random.randint(1,3)):
        sid+=1;q=random.choice(list(vs));sim=round(SequenceMatcher(None,norm(q),norm(n)).ratio(),3);ok=sim>=.92
        searches.append({'search_id':f'SID{sid:07d}','patient_id':p['patient_id'],'query_name':q,'search_method':random.choice(['Exact','Normalized','Fuzzy','Manual']),'similarity_score':sim,'phone_last4_match':random.random()<.94,'dob_match':random.random()<.96,'first_search_success':ok,'search_time_sec':round(random.uniform(8,25) if ok else random.uniform(60,420),1),'manual_intervention':not ok})

# Encounters + workflow
encounters=[];orders=[];tasks=[];control=[];routes=[];queues=[];billing=[];reports=[];journey=[]
start=datetime(2026,7,1,7);on=tn=qn=rn=bn=0
for e in range(1,5001):
    p=random.choice(patients);eid=f'ENC{e:07d}';arr=start+timedelta(minutes=random.randint(0,60*24*45));dept=random.choices(['Medicine OPD','Cardiology'],[.72,.28])[0];nsvc=random.choices([1,2,3,4,5,6],[.22,.26,.23,.16,.09,.04])[0];svcs=random.sample(list(services),nsvc);reg0=random.uniform(4,10);reg1=random.uniform(.5,2.5) if p['preferred_identity_token']!='Manual ID' else random.uniform(2.5,5)
    encounters.append({'encounter_id':eid,'patient_id':p['patient_id'],'arrival_ts':arr.isoformat(timespec='minutes'),'doctor_department':dept,'identity_method':p['preferred_identity_token'],'ordered_services':nsvc,'status':'Completed'})
    cur=dept;tw=walk=pay=over=0;ready=[]
    for seq,svc in enumerate(svcs,1):
        on+=1;tn+=1;qn+=1;rn+=1;bn+=1;oid=f'ORD{on:08d}';tid=f'TSK{tn:08d}';cat,price,stime=services[svc];cands=['Ultrasound'] if svc=='Ultrasound Abdomen' else sloc[cat]
        opts=[]
        for loc in cands:
            wt=random.randint(2,28);di=dist(cur,loc);score=wt+di/75+stime;opts.append((score,loc,wt,di))
        opts.sort();best=opts[0];override=random.random()<.08;chosen=random.choice(opts) if override and len(opts)>1 else best;over+=int(override);priority=random.choices(['Routine','Urgent'],[.9,.1])[0];ots=arr+timedelta(minutes=reg0+random.uniform(10,35))
        orders.append({'order_id':oid,'encounter_id':eid,'patient_id':p['patient_id'],'sequence':seq,'service_name':svc,'service_category':cat,'clinical_priority':priority,'ordered_ts':ots.isoformat(timespec='minutes'),'order_status':'Completed'})
        tasks.append({'task_id':tid,'order_id':oid,'encounter_id':eid,'task_type':'Perform Service','task_status':'Completed','assigned_location':chosen[1],'priority':priority,'routing_method':'Manual Override' if override else 'Rules+Optimization'})
        control.append({'task_id':tid,'encounter_id':eid,'displayed_to_front_desk':True,'front_desk_action':'Override' if override else 'Accept','action_delay_sec':round(random.uniform(3,18),1),'exception_reason':random.choice(['Capacity concern','Patient accessibility','Machine availability','Staff judgment']) if override else ''})
        qw=max(1,chosen[2]+random.randint(-2,5));queues.append({'queue_event_id':f'QE{qn:08d}','task_id':tid,'encounter_id':eid,'location':chosen[1],'queue_wait_min':qw,'service_time_min':stime+random.randint(-2,4),'sla_breach':qw>20})
        routes.append({'routing_id':f'RT{rn:08d}','task_id':tid,'encounter_id':eid,'from_location':cur,'recommended_location':best[1],'selected_location':chosen[1],'recommended_score':round(best[0],2),'selected_score':round(chosen[0],2),'predicted_wait_min':best[2],'distance_m':chosen[3],'human_override':override})
        disc=random.choice([0,0,0,50,100,150]);due=price-disc;billing.append({'ledger_id':f'BL{bn:08d}','encounter_id':eid,'order_id':oid,'service_name':svc,'department':chosen[1],'gross_charge_bdt':price,'discount_bdt':disc,'covered_bdt':0,'patient_payable_bdt':due,'billing_mode':'Encounter Ledger','settlement_status':'Settled'})
        rt=ots+timedelta(minutes=qw+stime+random.randint(15,180));reports.append({'encounter_id':eid,'order_id':oid,'service_name':svc,'result_status':'Ready','ready_ts':rt.isoformat(timespec='minutes'),'delivery_channel':random.choice(['Unified Report Hub','Portal','Printed at Hub'])});ready.append(rt);tw+=qw;walk+=chosen[3];pay+=due;cur=chosen[1]
    base_admin=reg0+nsvc*random.uniform(3.2,5.8)+random.uniform(7,16);new_admin=reg1+random.uniform(4,10);base_walk=walk+nsvc*random.randint(90,180);new_walk=max(100,walk-random.randint(80,240))+random.randint(40,100)
    journey.append({'encounter_id':eid,'patient_id':p['patient_id'],'service_count':nsvc,'baseline_registration_min':round(reg0,2),'digital_registration_min':round(reg1,2),'baseline_payment_touchpoints':nsvc,'centralized_payment_touchpoints':1,'baseline_admin_min':round(base_admin,2),'redesigned_admin_min':round(new_admin,2),'baseline_walking_m':base_walk,'redesigned_walking_m':new_walk,'service_queue_min':tw,'human_routing_overrides':over,'total_payable_bdt':pay,'all_reports_ready_ts':max(ready).isoformat(timespec='minutes')})

# HR + financial + predictive
staff=[];training=[];job=[]
for i in range(1,181):
    emp=f'EMP{i:04d}';role=random.choice(['Front Desk Officer','Billing Officer','Lab Coordinator','Patient Navigator','Operations Executive']);training.append({'employee_id':emp,'role':role,'digital_training_hours':random.choice([4,6,8,12]),'assessment_score':random.randint(65,98),'adoption_status':random.choices(['Adopted','Needs Coaching','Not Yet Adopted'],[.78,.18,.04])[0]});before=random.randint(45,75);after=random.randint(8,25);job.append({'employee_id':emp,'role':role,'manual_admin_share_before_pct':before,'manual_admin_share_after_pct':after,'patient_support_share_before_pct':random.randint(10,25),'patient_support_share_after_pct':random.randint(25,45),'exception_management_share_after_pct':random.randint(15,35)})
for day in range(45):
    d=(datetime(2026,7,1)+timedelta(days=day)).date().isoformat()
    for loc in ['Reception','Central Billing','Lab A','Lab B','Radiology A','Radiology B','ECG','Ultrasound','Report Hub']:
        for shift in ['Morning','Evening']:
            sched=random.randint(2,7);active=max(1,sched-random.choice([0,0,0,1]));demand=random.randint(40,180);staff.append({'date':d,'location':loc,'shift':shift,'scheduled_staff':sched,'active_staff':active,'patient_demand':demand,'patients_per_active_staff':round(demand/active,2),'overtime_hours':round(max(0,(demand-active*24)/18),2),'absence_flag':active<sched})
avg_saved=sum(r['baseline_admin_min']-r['redesigned_admin_min'] for r in journey)/len(journey);annual=220000;hours=annual*avg_saved/60;sc=[('A','Add 10 FTE',5200000,2600000,.20,.10,.08),('B','Workflow redesign only',2500000,700000,.38,.35,.45),('C','Digital identity + routing + central billing',6800000,1600000,.58,.55,.62),('D','Hybrid: digital + 4 targeted FTE',8200000,2600000,.70,.64,.72)]
financial=[]
for code,name,impl,op,wait,admin,cap in sc:
    lv=hours*280*admin;ov=hours*420*cap;benefit=lv+ov;financial.append({'scenario_code':code,'scenario_name':name,'implementation_cost_bdt':impl,'annual_operating_cost_bdt':op,'wait_reduction_pct':wait,'admin_time_reduction_pct':admin,'capacity_gain_pct':cap,'annual_labor_capacity_value_bdt':round(lv,2),'annual_operational_value_bdt':round(ov,2),'annual_total_benefit_bdt':round(benefit,2),'year1_net_benefit_bdt':round(benefit-impl-op,2),'simple_payback_years':round(impl/max(1,benefit-op),2)})
predict=[]
for i,q in enumerate(queues,1):
    hour=random.randint(7,21);active=random.randint(1,6);demand=random.randint(10,80);recent=max(1,q['queue_wait_min']+random.randint(-6,6));pred=max(1,.6*recent+.5*(demand/active)+(4 if 9<=hour<=11 else 0));predict.append({'record_id':f'PR{i:08d}','location':q['location'],'hour':hour,'day_of_week':random.randint(0,6),'active_staff':active,'arrivals_last_30m':demand,'recent_avg_wait_min':recent,'predicted_wait_min':round(pred,2),'predicted_sla_breach':pred>20,'actual_wait_min':q['queue_wait_min']})
kpis=[{'encounter_id':r['encounter_id'],'admin_minutes_saved':round(r['baseline_admin_min']-r['redesigned_admin_min'],2),'payment_touchpoints_saved':r['baseline_payment_touchpoints']-1,'walking_m_saved':r['baseline_walking_m']-r['redesigned_walking_m'],'service_queue_min':r['service_queue_min'],'routing_overrides':r['human_routing_overrides'],'total_payable_bdt':r['total_payable_bdt']} for r in journey]

# Write datasets
for rel,rows in [('data/01_raw_synthetic/patients.csv',patients),('data/01_raw_synthetic/encounters.csv',encounters),('data/02_identity_resolution/identity_name_variants.csv',variants),('data/02_identity_resolution/identity_search_events.csv',searches),('data/03_workflow/service_orders.csv',orders),('data/03_workflow/workflow_tasks.csv',tasks),('data/03_workflow/front_desk_control_tower.csv',control),('data/04_routing/hospital_nodes.csv',nodes),('data/04_routing/hospital_edges.csv',edges),('data/04_routing/routing_decisions.csv',routes),('data/04_routing/queue_events.csv',queues),('data/05_billing_reports/billing_ledger.csv',billing),('data/05_billing_reports/report_status.csv',reports),('data/06_hr_operations/staff_shifts.csv',staff),('data/06_hr_operations/training_adoption.csv',training),('data/06_hr_operations/job_redesign.csv',job),('data/07_financial_model/intervention_scenarios.csv',financial),('data/08_predictive_operations/predictive_queue_features.csv',predict),('data/09_analysis_ready/journey_kpis.csv',kpis),('data/09_analysis_ready/journey_summary.csv',journey)]:write_csv(rel,rows)

# Data dictionary
dictionary=[]
for p in (ROOT/'data').rglob('*.csv'):
    with p.open(encoding='utf-8-sig') as f:r=csv.DictReader(f);firstrow=next(r)
    for field,val in firstrow.items():dictionary.append({'table':str(p.relative_to(ROOT)),'field':field,'data_type':'number' if str(val).replace('.','',1).isdigit() else 'text','description':field.replace('_',' ').title(),'synthetic_data':'Yes','privacy_note':'No real patient/employee identity in public dataset'})
write_csv('data_dictionary/data_dictionary.csv',dictionary)

# Source modules
(ROOT/'src/identity.py').write_text("from difflib import SequenceMatcher\ndef normalize_name(s): return ' '.join(s.lower().replace('.','').split())\ndef identity_score(a,b): return round(SequenceMatcher(None,normalize_name(a),normalize_name(b)).ratio(),4)\n",encoding='utf-8')
(ROOT/'src/routing.py').write_text("def route_score(wait_min,distance_m,service_time_min,capacity_penalty=0): return wait_min+distance_m/75+service_time_min+capacity_penalty\n",encoding='utf-8')
(ROOT/'src/workflow.py').write_text("VALID=['ordered','ready','accepted','in_progress','completed','cancelled','exception']\ndef transition(task,state):\n    if state not in VALID: raise ValueError('invalid state')\n    task=dict(task);task['task_status']=state;return task\n",encoding='utf-8')
(ROOT/'src/billing.py').write_text("def encounter_total(rows): return sum(float(r['patient_payable_bdt']) for r in rows)\ndef payment_touchpoints(service_count,centralized=True): return 1 if centralized else max(1,service_count)\n",encoding='utf-8')
(ROOT/'src/predictive.py').write_text("def predict_wait(recent,arrivals,staff,peak=False): return round(max(1,.6*recent+.5*(arrivals/max(1,staff))+(4 if peak else 0)),2)\n",encoding='utf-8')

# Notebooks
notebook_specs=[('01_identity_resolution.ipynb','Identity Resolution','Measure lookup and manual intervention',"import csv\nrows=list(csv.DictReader(open('../data/02_identity_resolution/identity_search_events.csv',encoding='utf-8-sig')))\nprint(len(rows))"),('02_control_tower.ipynb','Control Tower','Analyze accept/override behavior',"import csv\nrows=list(csv.DictReader(open('../data/03_workflow/front_desk_control_tower.csv',encoding='utf-8-sig')))\nprint(sum(r['front_desk_action']=='Override' for r in rows)/len(rows))"),('03_routing_navigation.ipynb','Routing & Navigation','Evaluate queue-aware routing',"import csv\nrows=list(csv.DictReader(open('../data/04_routing/routing_decisions.csv',encoding='utf-8-sig')))\nprint(len(rows))"),('04_central_billing.ipynb','Central Billing','Compare payment touchpoints',"import csv,statistics\nrows=list(csv.DictReader(open('../data/09_analysis_ready/journey_summary.csv',encoding='utf-8-sig')))\nprint(statistics.mean(int(r['baseline_payment_touchpoints']) for r in rows))"),('05_hr_operations.ipynb','HR Operations','Analyze staffing demand',"import csv,statistics\nrows=list(csv.DictReader(open('../data/06_hr_operations/staff_shifts.csv',encoding='utf-8-sig')))\nprint(statistics.mean(float(r['patients_per_active_staff']) for r in rows))"),('06_financial_model.ipynb','Financial Model','Compare interventions',"import csv\nrows=list(csv.DictReader(open('../data/07_financial_model/intervention_scenarios.csv',encoding='utf-8-sig')))\nprint(max(rows,key=lambda r:float(r['year1_net_benefit_bdt'])))"),('07_predictive_ops.ipynb','Predictive Operations','Queue forecast baseline',"import csv\nrows=list(csv.DictReader(open('../data/08_predictive_operations/predictive_queue_features.csv',encoding='utf-8-sig')))\nprint(len(rows))"),('08_challenge_baseline.ipynb','Challenge Baseline','Starter learner analysis',"import csv\nrows=list(csv.DictReader(open('../data/09_analysis_ready/journey_kpis.csv',encoding='utf-8-sig')))\nprint(len(rows))"),('09_executive_recommendation.ipynb','Executive Recommendation','Combine people, process and technology evidence',"import csv\nrows=list(csv.DictReader(open('../data/07_financial_model/intervention_scenarios.csv',encoding='utf-8-sig')))\nprint(max(rows,key=lambda r:float(r['year1_net_benefit_bdt']))['scenario_name'])")]
for fn,t,d,c in notebook_specs:(ROOT/'notebooks'/fn).write_text(json.dumps(nb(t,d,c),indent=2),encoding='utf-8')

# Docs
(ROOT/'README.md').write_text("""# BD Patient Journey Analytics\n\n**v1.0.0 — Bangladesh National Medical City (BNMC), synthetic demo**\n\nA portfolio-ready **HR Operations × HRBP × Healthcare Operations × IT × Finance × Data Analytics** case study.\n\n## Core case\nAfter a doctor creates service orders, the hospital workflow converts them into operational tasks, sends them to a front-desk control tower, routes the patient to the right service point, consolidates billing into one encounter ledger, and centralizes report collection.\n\n**Doctor → Service Orders → Front Desk Control Tower → Smart Routing → Central Billing → Report Hub & Exit**\n\n## What v1.0 includes\n- synthetic identity-resolution and patient-flow datasets\n- doctor order → task workflow and human override logging\n- hospital graph, queue-aware routing and walking-distance analysis\n- centralized billing and report-hub model\n- staffing, training, job-redesign and HR Operations datasets\n- people vs process vs technology vs hybrid financial scenarios\n- predictive queue/SLA baseline\n- Excel executive workbook + 9 Python notebooks + learner challenge\n\n## Governance\nAll records are synthetic. Operational AI supports routing/forecasting only; clinicians retain clinical decisions. Do not add real patient-identifying or biometric data to public submissions.\n""",encoding='utf-8')
(ROOT/'PROJECT_CHARTER.md').write_text('# Project Charter\n\nDecision question: should management solve post-doctor patient-flow friction through staffing, process redesign, digital workflow, centralized billing/reporting, or a hybrid? HR Operations/HRBP evaluates workforce planning, job design, training, adoption, productivity and ROI.\n',encoding='utf-8')
(ROOT/'ROADMAP.md').write_text('# Version Roadmap\n\n- ✅ v0.1 Foundation\n- ✅ v0.2 Identity resolution\n- ✅ v0.3 Control tower\n- ✅ v0.4 Routing/navigation\n- ✅ v0.5 Central billing/reports\n- ✅ v0.6 HR Operations\n- ✅ v0.7 Financial modelling\n- ✅ v0.8 Predictive operations\n- ✅ v0.9 Community beta\n- ✅ v1.0 Portfolio release\n',encoding='utf-8')
(ROOT/'DATA_PROVENANCE.md').write_text('# Data Provenance\n\nAll records are programmatically generated synthetic data. No real hospital patient, biometric, billing or employee records are included.\n',encoding='utf-8')
(ROOT/'DATASET_USAGE_GUIDE.md').write_text('# Dataset Usage Guide\n\nStart with `data/09_analysis_ready/journey_summary.csv`, then identity, workflow, routing, billing, HR, financial and predictive tables. Use the Excel workbook for management/HRBP analysis and notebooks for Python practice.\n',encoding='utf-8')
(ROOT/'SECURITY.md').write_text('# Security & Privacy\n\nNever commit credentials, real patient-identifying data or real biometric templates. Public submissions must remain synthetic/de-identified. Human override and auditability are required for operational AI recommendations.\n',encoding='utf-8')
mit='''MIT License\n\nCopyright (c) 2026 Musa\n\nPermission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the "Software"), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:\n\nThe above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.\n\nTHE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.\n'''
(ROOT/'LICENSE').write_text(mit,encoding='utf-8');(ROOT/'DATA_LICENSE.md').write_text('# Synthetic Data License\n\nSynthetic datasets in `data/` are released under Creative Commons Attribution 4.0 International (CC BY 4.0).\n',encoding='utf-8')
(ROOT/'CITATION.cff').write_text('cff-version: 1.2.0\ntitle: "BD Patient Journey Analytics"\nmessage: "Please cite this project if reused."\nversion: 1.0.0\ndate-released: 2026-08-21\nauthors:\n  - family-names: "Musa"\n    given-names: "Musa"\nlicense: MIT\n',encoding='utf-8')
(ROOT/'CHANGELOG.md').write_text('# Changelog\n\n## v1.0.0 — 2026-08-21\nIntegrated identity, workflow, routing, billing, HR Operations, financial modelling, predictive operations and learner challenge.\n',encoding='utf-8')
(ROOT/'releases/v1.0.0.md').write_text('# Release v1.0.0\n\nPortfolio-ready integrated release for the synthetic BNMC post-doctor patient journey case study. Includes reproducible datasets, Excel workbook, Python notebooks, governance documentation and learner challenge assets.\n',encoding='utf-8')
for ver in ['v0.1.0','v0.2.0','v0.3.0','v0.4.0','v0.5.0','v0.6.0','v0.7.0','v0.8.0','v0.9.0']:(ROOT/'releases'/f'{ver}.md').write_text(f'# Release {ver}\n\nMilestone completed in the v1.0 integrated build.\n',encoding='utf-8')
(ROOT/'case_study/problem_statement.md').write_text('# Case Study\n\nDesign the lowest-friction post-doctor patient journey without compromising clinical priority, billing control, privacy or operational accountability.\n',encoding='utf-8')
(ROOT/'case_study/evaluation_rubric.md').write_text('# Evaluation Rubric — 100\n\nProblem 10; data 10; workflow/routing 20; billing 10; HR Operations 15; financial model 15; visualization 10; recommendation/limitations 10.\n',encoding='utf-8')
(ROOT/'competition/competition_brief.md').write_text('# Community Challenge\n\nTracks: identity resolution, queue/routing, HR Operations, central billing, and integrated HRBP business case.\n',encoding='utf-8')
(ROOT/'competition/scoring.md').write_text('# Scoring\n\n30% patient-time improvement; 20% feasibility; 15% workforce; 15% finance; 10% governance; 10% reproducibility.\n',encoding='utf-8')
for rel,text in [('assignments/01_beginner_excel/README.md','# Beginner Excel\nBuild PivotTables and dashboards for admin time, payment touchpoints, walking distance and queues.\n'),('assignments/02_intermediate_python/README.md','# Intermediate Python\nAnalyze identity, workflow exceptions, routing, staffing and billing.\n'),('assignments/03_advanced_optimization/README.md','# Advanced Optimization\nBuild match scoring, route optimization, staffing scenarios and SLA prediction.\n'),('assignments/04_hrbp_business_case/README.md','# HRBP Business Case\nRecommend staffing, process, technology or hybrid using service, workforce, cost, risk and change-management evidence.\n')]: (ROOT/rel).write_text(text,encoding='utf-8')
(ROOT/'.github/ISSUE_TEMPLATE/case-study-submission.yml').write_text('name: Case Study Submission\ndescription: Submit a reproducible solution\ntitle: "[CASE] "\nbody:\n  - type: textarea\n    id: summary\n    attributes:\n      label: Executive summary\n    validations:\n      required: true\n  - type: checkboxes\n    id: privacy\n    attributes:\n      label: Privacy\n      options:\n        - label: I used synthetic/de-identified data only.\n          required: true\n',encoding='utf-8')
(ROOT/'.github/DISCUSSION_TEMPLATE/solution-idea.yml').write_text('title: "[IDEA] "\nbody:\n  - type: textarea\n    id: solution\n    attributes:\n      label: Proposed solution\n    validations:\n      required: true\n',encoding='utf-8')
(ROOT/'kaggle/dataset-metadata.json').write_text(json.dumps({'title':'BD Patient Journey Analytics - Synthetic','id':'REPLACE_KAGGLE_USERNAME/bd-patient-journey-analytics','licenses':[{'name':'CC-BY-4.0'}],'keywords':['healthcare','hr-analytics','operations','synthetic-data','patient-flow','queue','excel','python']},indent=2),encoding='utf-8')
(ROOT/'kaggle/README.md').write_text('# Kaggle Distribution\n\nPublish only after replacing the Kaggle username and verifying credentials. Use the clean-package builder.\n',encoding='utf-8')

# Excel workbook
wb=Workbook();ws=wb.active;ws.title='Executive Dashboard';js=wb.create_sheet('Journey Summary');fs=wb.create_sheet('Financial Scenarios');hrs=wb.create_sheet('HR Staffing');ids=wb.create_sheet('Identity')
blue='1F4E78';dark='17365D';white='FFFFFF'
def header(row):
    for c in row:c.fill=PatternFill('solid',fgColor=blue);c.font=Font(color=white,bold=True);c.alignment=Alignment(horizontal='center')
# journey sample
jh=['encounter_id','service_count','baseline_payment_touchpoints','centralized_payment_touchpoints','baseline_admin_min','redesigned_admin_min','baseline_walking_m','redesigned_walking_m','service_queue_min','routing_overrides','total_payable_bdt'];js.append(jh);header(js[1])
for r in journey[:1500]:js.append([r['encounter_id'],r['service_count'],r['baseline_payment_touchpoints'],1,r['baseline_admin_min'],r['redesigned_admin_min'],r['baseline_walking_m'],r['redesigned_walking_m'],r['service_queue_min'],r['human_routing_overrides'],r['total_payable_bdt']])
fh=list(financial[0]);fs.append(fh);header(fs[1]);[fs.append([r[h] for h in fh]) for r in financial]
hh=list(staff[0]);hrs.append(hh);header(hrs[1]);[hrs.append([r[h] for h in hh]) for r in staff]
ih=['search_id','search_method','similarity_score','search_time_sec','success_flag','manual_flag'];ids.append(ih);header(ids[1]);[ids.append([r['search_id'],r['search_method'],r['similarity_score'],r['search_time_sec'],int(r['first_search_success']),int(r['manual_intervention'])]) for r in searches[:1500]]
ws.merge_cells('A1:H2');ws['A1']='BD Patient Journey Analytics — v1.0.0';ws['A1'].fill=PatternFill('solid',fgColor=dark);ws['A1'].font=Font(color=white,bold=True,size=18);ws['A1'].alignment=Alignment(horizontal='center')
metrics=[('Avg baseline admin min',"=AVERAGE('Journey Summary'!E2:E1501)"),('Avg redesigned admin min',"=AVERAGE('Journey Summary'!F2:F1501)"),('Avg admin min saved','=B5-B6'),('Avg baseline payment touches',"=AVERAGE('Journey Summary'!C2:C1501)"),('Avg centralized payment touches',"=AVERAGE('Journey Summary'!D2:D1501)"),('Avg baseline walking m',"=AVERAGE('Journey Summary'!G2:G1501)"),('Avg redesigned walking m',"=AVERAGE('Journey Summary'!H2:H1501)"),('Avg walking m saved','=B10-B11')]
ws.append([]);ws.append(['Patient Journey KPI','Value']);header(ws[4])
for name,formula in metrics:ws.append([name,formula])
ws['D4']='Governance';ws['E4']='Rule';header([ws['D4'],ws['E4']]);gov=[('Clinical decisions','Clinician-owned'),('Operational AI','Routing/forecast only'),('Human override','Required'),('Public data','Synthetic only'),('Biometric data','Not included'),('Billing','Encounter ledger'),('Front desk','Control tower')]
for i,(a,b) in enumerate(gov,5):ws.cell(i,4,a);ws.cell(i,5,b)
ws['G4']='Intervention';ws['H4']='Year-1 Net Benefit';header([ws['G4'],ws['H4']])
for i,r in enumerate(financial,5):ws.cell(i,7,r['scenario_code']);ws.cell(i,8,r['year1_net_benefit_bdt'])
chart=BarChart();chart.title='Administrative Minutes per Encounter';chart.add_data(Reference(ws,min_col=2,min_row=5,max_row=6),titles_from_data=False);chart.set_categories(Reference(ws,min_col=1,min_row=5,max_row=6));ws.add_chart(chart,'J4')
for sh in wb.worksheets:
    sh.freeze_panes='A2'
    for idx,col in enumerate(sh.iter_cols(),1):
        letter=get_column_letter(idx)
        vals=list(col)[:100]
        sh.column_dimensions[letter].width=min(28,max(12,max(len(str(c.value or '')) for c in vals)+2))
wb.save(ROOT/'excel/Hospital_Digital_Patient_Journey_HR_Operations_v1.0.xlsx')

manifest={'project':'BD Patient Journey Analytics','demo_hospital':'Bangladesh National Medical City (BNMC)','version':'1.0.0','release_date':'2026-08-21','synthetic_only':True,'counts':{'patients':len(patients),'identity_search_events':len(searches),'encounters':len(encounters),'service_orders':len(orders),'workflow_tasks':len(tasks),'routing_decisions':len(routes),'billing_rows':len(billing),'staff_shift_rows':len(staff),'predictive_rows':len(predict)}}
(ROOT/'MANIFEST.json').write_text(json.dumps(manifest,indent=2),encoding='utf-8')
(ROOT/'requirements.txt').write_text('openpyxl>=3.1,<4\n',encoding='utf-8')
print(json.dumps(manifest,indent=2))
